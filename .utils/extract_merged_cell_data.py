# extract_merged_cell_data.py 20250826

import openpyxl

def extract_merged_cell_data(file_path):
    """
    Extracts and prints string data from merged cells in an Excel file.
    
    Args:
        file_path (str): The path to the Excel file.
    """
    try:
        # Load the workbook from the specified file path
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: The file at {file_path} was not found.")
        return

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"\n--- Processing Sheet: '{sheet_name}' ---")

        # Get the list of merged cell ranges for the current sheet
        merged_cells = sheet.merged_cells.ranges

        # Iterate through each merged cell range
        for merged_range in merged_cells:
            # The value of a merged cell is stored in the top-left cell of the range.
            top_left_cell = merged_range.min_row, merged_range.min_col
            cell = sheet.cell(row=top_left_cell[0], column=top_left_cell[1])
            
            # Check if the cell contains a string value and print it
            if isinstance(cell.value, str):
                print(cell.value)

# Specify the path to your Excel file
excel_file_path = "Monthly Calendar INTERNAL 2025.xlsx"  # <--- Change this to your file's path
extract_merged_cell_data(excel_file_path)