# gen_court_session_calendar.py 20251021

import calendar
import configparser
from datetime import date, datetime, timedelta
import duckdb
from enum import Enum
from icecream import ic
from loguru import logger
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Side, Fill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
import os
import pandas as pd
from pandas import DataFrame # Import the specific class for hinting
from pathlib import Path
import pyodbc
import re
from sqlalchemy import create_engine
import sys
import urllib
import typer
import yaml

#==============================================================================
# Constants
#==============================================================================
debugging_skip_code = False
MAX_ROW = None
#------------------------------------------------------------------------------
#==============================================================================
# Classes
#==============================================================================
class SourcePDFError(Exception):
    """The source PDF causes an error."""
    pass
#------------------------------------------------------------------------------
#==============================================================================
# Functions
#==============================================================================
def setup_logging(config: configparser.ConfigParser):
    logs_folder = config['Paths']['logs_folder']
    # console logging
    console_log_level = config['Logging']['console_level']
    # Define a custom format string with color tags for the console output
    console_format = "<light-green>{time:YYYY-MM-DD HH:mm:ss}</light-green> | <level>{level: <8}</level> | <bold><cyan>{function:<20}</cyan>:<cyan>{line:>5}</cyan></bold> - <level>{message}</level>"
    console_format = config['Logging'].get('console_format',console_format)
    # file logging
    file_log_level = config['Logging']['file_level']
    # Define a custom log file path with a dynamic date
    file_log_path = Path(logs_folder) / Path(f"{Path(__file__).stem}"+"_{time:YYYY-MM-DD}.log")
    # Define a simple format string for the log file (no colors needed)
    file_format = "{time:YYYY-MM-DD HH:mm:ss} | {level:<8} | {function:<20}:{line:>5} - {message}"
    file_format = config['Logging'].get('file_format',file_format)
    # Remove the default handler
    logger.remove(0)
    # Add a sink for the console with the colored format
    logger.add(
        sys.stderr, # Use sys.stderr for console output
        level=console_log_level.upper(), 
        format=console_format,
        backtrace=False,
        diagnose=False,
        enqueue=True
    )
    # Add a sink for the log file with the plain format, daily rotation, and retention
    logger.add(
        file_log_path, 
        level=file_log_level.upper(), 
        rotation="1 day", 
        retention="1 week",
        format=file_format,
        backtrace=True,
        diagnose=True,
        enqueue=True
    )
    return logger
# =================================================================================================
# Read applicaytion configuration from a YAML file.
# =================================================================================================
def read_yaml_configuration(yamlFilename: str) -> configparser:
    """
    Reads the applicaiton's configuration informaiton from the given filename, yamlFilename. 
    Returns: A tuple (
        dfFinCols: Dictionary containing the Finances DataFrame structure, 
        # This is definition of the DataFrame, with column names and types, that the input CSV data 
        # files will be converted to.
        sourcedata: An array containg a dictionary for each type of input data file
        # Each dictionary defines how to convert its corresponding CSV data file into the finances DataFrame.
    )
    """
    ymldata = None
    with open(yamlFilename, 'r') as file:
        ymldata = yaml.safe_load(file)
    return(ymldata)
# ------------------------------------------------------------------------------------------------+
def get_yaml_config_color(yaml_conf: configparser, color_name: str) -> str:
    return yaml_conf['constants']['colors'][color_name]
# ------------------------------------------------------------------------------------------------+
def date_range_generator(start_date: date, end_date: date):
    """Generates a sequence of dates from start_date up to and including end_date."""
    # 1. Calculate the duration (a timedelta object)
    delta = end_date - start_date
    # 2. Iterate using range() for the number of days in the duration
    # delta.days gives the number of days, +1 makes the range inclusive of the end_date.
    for i in range(delta.days + 1):
        # 3. Yield the start_date plus the timedelta for the current iteration (i days)
        yield start_date + timedelta(days=i)
# ------------------------------------------------------------------------------------------------+
def get_date_range(start_date: date, end_date: date):
    """Generates a list of dates between a start and end date (inclusive)."""
    date_list = []
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)
    return date_list
#--------------------------------------------------------------------------------------------------
def find_first_matching_cell(worksheet: Worksheet, column_letter: str, text: str) -> int:
    """
    Finds the first row index (1-based) with the given text in the given column.
    Returns the row number of the matching cell or None.
    """
    for row_num in range(1,worksheet.max_row+2):
        cell = worksheet[f'{column_letter}{row_num}']
        if cell.value == text or cell.value is None and (text == '' or text is None): #if cell.value == text:
            return row_num
    logger.debug(f"No empty cell found in column {column_letter} up to row {worksheet.max_row + 1}.")
    return None
#--------------------------------------------------------------------------------------------------
def find_first_matching_cell_by_col_idx(worksheet: Worksheet, column_idx: int, text: str,start_row: int = 1) -> int:
    """
    Finds the first row index (1-based) with the given text in the given column.
    Returns the row number of the matching cell or None.
    """
    for row_num in range(start_row,worksheet.max_row+2):
        cell = worksheet.cell(row_num,column_idx)
        if cell.value == text or cell.value is None and (text == '' or text is None): #if cell.value == text:
            return row_num
    logger.debug(f"Cell with value '{text}' not found in column index {column_idx} up to row {worksheet.max_row + 1}.")
    return None
#--------------------------------------------------------------------------------------------------
def get_weekday_number(year: int, month: int, day: int):
    """
    Returns the weekday as an integer (1=Monday, 5=Sunday).
    """
    try:
        # Create a date object
        d = date(year, month, day)
        
        # Use the .weekday() method
        return d.isoweekday()
    except ValueError as e:
        # Handles errors like an invalid date (e.g., February 30th)
        logger.exception(f"Error creating date: {year}-{month}-{day}, {e}")
        return None
#--------------------------------------------------------------------------------------------------
def copy_cell(source_cell: Cell, dest_cell: Cell):
    """
    Copies the value, number format, and style properties of one cell to another.
    """
    # --- 1. Copy Value and Type ---
    dest_cell.value = source_cell.value
    # --- 2. Copy Number Formatting ---
    # This handles currency, date formats, percentages, etc.
    dest_cell.number_format = source_cell.number_format
    # --- 3. Copy Basic Styles (Font, Fill, Border, Alignment, etc.) ---
    # Note: openpyxl stores styles internally. Accessing the '_style' attribute 
    # provides the most reliable way to copy the entire style object.
    if source_cell.has_style:
         dest_cell._style = source_cell._style
    # --- 4. Copy Cell Protection (Locked/Hidden) ---
#    if source_cell.protection:
#        dest_cell.protection = source_cell.protection
#    # --- 5. Copy Comment (if present) ---
    if source_cell.comment:
        dest_cell.comment = source_cell.comment
#--------------------------------------------------------------------------------------------------
def get_odyssey_court_sessions_by_year(year: int,config: configparser) -> DataFrame:
    # Construct the ODBC connection string with Trusted_Connection=yes
    odbc_conn_str = (
        f"DRIVER={{{config['SQL']['driver_name']}}};"  # Note the double curly braces for the driver name
        f"SERVER={config['SQL']['server_name']};"
        f"DATABASE={config['SQL']['database_name']};"
        f"Trusted_Connection=yes;"
    )
    # URL-encode the entire ODBC connection string
    params = urllib.parse.quote_plus(odbc_conn_str)
    # Create the Database URI using the mssql+pyodbc dialect
    # The format is: dialect+driver:///?odbc_connect=params
    DB_URI = f"mssql+pyodbc:///?odbc_connect={params}"
    # Create the SQLAlchemy Engine
    engine = create_engine(DB_URI)
    logger.info("SQLAlchemy Engine created successfully using Windows Authentication.")
    # Define SQL stament to execute the stored procedure.
    sql_query = f"""
        SET NOCOUNT ON;
        exec Justice.fc.sp_getCourtSessionsByYear @pMonthOrYear=?;
    """
    # Execute the SQL stored procedure to get all of the Odyssey court sessions for the
    # given year and store the results in a Pandas DataFrame.
    try:
        df = pd.read_sql(
            sql=sql_query,
            con=engine,      # Pass the SQLAlchemy Engine
            params=(year,) 
        )
        df['SessionDate'] = pd.to_datetime(df['SessionDate'])
    except Exception as e:
        logger.exception(f"\nAn error occurred: {e}")
    finally:
        engine.dispose()
    return df
#--------------------------------------------------------------------------------------------------
def convert_df_to_list(df: DataFrame,yaml_config) -> list:
    court_session_list = []
    sessions = []
    try:
        # Create a connection to an in memory DuckDB DB.
        #ddb_conn = duckdb.connect(database=':memory:')
        ddb_conn = duckdb.connect(database=r'.zsandbox\court_calendar.db')
        #
        # Get the judges from the YAML configuration
        judges = []
        judges = yaml_config['data']['superior_judges'] + yaml_config['data']['state_judges']
        # Create table and insert the judges
        DUCKDB_TABLE_NAME = "judge"
        ddb_conn.sql(f"""
drop table if exists {DUCKDB_TABLE_NAME};
CREATE TABLE {DUCKDB_TABLE_NAME} (
    Name VARCHAR
    ,OysseyCode VARCHAR
    ,Color VARCHAR
 );
        """)
        # Insert the judges
        # DuckDB handles the list insertion automatically when passed as a parameter.
        for record in judges:
            ordered_values = (
                record['name']
                ,record['odyssey_code']
                ,record['color']
            )
            ddb_conn.execute(f"INSERT INTO {DUCKDB_TABLE_NAME} VALUES (?, ?, ?)", ordered_values)
        #
        #
        # Get the Special Dates from the YAML configuration
        special_dates = []
        #breakpoint()
        for sp_dt in yaml_config['data']['special_dates']:
            if sp_dt.get('date',None):
                special_dates.append(sp_dt)
            else:
                for dt in date_range_generator(sp_dt['begin_date'],sp_dt['end_date']):
                    new_sp_dt = {
                        'name': sp_dt['name']
                        ,'date': dt
                        ,'color': sp_dt['color']
                        ,'display_order': sp_dt['display_order']
                    }
                    special_dates.append(new_sp_dt)
        # Create table and insert the special dates.
        DUCKDB_TABLE_NAME = "special_date"
        ddb_conn.sql(f"""
drop table if exists {DUCKDB_TABLE_NAME};
CREATE TABLE {DUCKDB_TABLE_NAME} (
    Name VARCHAR
    ,Date DATE
    ,Color VARCHAR
    ,DisplayOrder int
 );
        """)
        # Insert the special_dates.
        # DuckDB handles the list insertion automatically when passed as a parameter.
        for record in special_dates:
            ordered_values = (
                record['name'] 
                ,record['date'] 
                ,record['color']
                ,record['display_order']
            )
            ddb_conn.execute(f"INSERT INTO {DUCKDB_TABLE_NAME} VALUES (?, ?, ?, ?)", ordered_values)
        #
        # Get the Court Session mappings from the YAML configuration.
        courtsession_mapping_list = yaml_config['data']['sessions_mapping']
        # Create the table schema first
        DUCKDB_TABLE_NAME = "courtsession_mapping"
        ddb_conn.sql(f"""
drop table if exists {DUCKDB_TABLE_NAME};
CREATE TABLE {DUCKDB_TABLE_NAME} (
    OdysseyCourtSession VARCHAR
    ,CalendarFormat VARCHAR
    ,DisplayOrder int
);
        """)
        # Insert the entire list of records using the built-in VALUES clause
        # DuckDB handles the list insertion automatically when passed as a parameter.
        for record in courtsession_mapping_list:
            ordered_values = (
                record['odyssey_name'] 
                ,record['calendar_name'] 
                ,record['display_order']
            )
            ddb_conn.execute(f"INSERT INTO {DUCKDB_TABLE_NAME} VALUES (?, ?, ?)", ordered_values)
        #
        # Create Court Sessions DB table from DataFramedf.
        DUCKDB_TABLE_NAME = 'courtsession'
        ddb_conn.sql(f"""
            drop table if exists {DUCKDB_TABLE_NAME};
            CREATE TABLE {DUCKDB_TABLE_NAME} AS SELECT * FROM df;
        """)
        # Create a list of the court sessions transitioned via the mapping and special_dates
        # and resulting in a the form needed to be displayed in the calendar.
        sql_qry="""
-- ================================================================================================
drop table if exists tmp_courtsession;
CREATE TEMPORARY TABLE tmp_courtsession
(
  SessionDate date
  ,StartTime varchar
  ,SessionDescription varchar
  ,Color varchar
  ,JudicialOfficerCode varchar
  ,DisplayOrder int
  ,Week int
);
insert into
  tmp_courtsession
select
  strftime(SessionDate,'%Y-%m-%d') as SessionDate
  ,case 
    when cs_m.CalendarFormat is not null
    then ''
    else cs.StartTime
  end as StartTime
  ,case 
    when cs_m.CalendarFormat is not null
    then
      replace(
        replace(
          cs_m.CalendarFormat
          ,'$CourtRoom}$'
          ,cs.CourtRoomCode
        )
        ,'${JudicialOfficer}$'
        ,cs.JudicialOfficerCode
      )
    else 
      concat(
        strftime(
          concat(
            '2025-01-01 '
            ,cs.StartTime
          )::datetime
          ,'%-I:%M '
        )
        ,regexp_replace(
          cs.SessionDescription
          ,'\\([A-Z]{3}\\) '
          ,''
        )
        ,' ('
        ,left(reverse(cs.JudicialOfficerCode),1)
        ,')'
      )
  end as SessionDescription
  ,j.Color as Color
  ,cs.JudicialOfficerCode as JudicialOfficerCode
  ,if(cs_m.DisplayOrder is null,3,cs_m.DisplayOrder) as DisplayOrder
  ,strftime(SessionDate,'%U')::int as Week
from
  courtsession cs 
  left outer join judge j
  on
   j.OysseyCode = cs.JudicialOfficerCode
  left outer join courtsession_mapping cs_m 
  on
    cs.SessionDescription ilike concat(
      '%'
      ,replace(
        replace(
          cs_m.OdysseyCourtSession
          ,'-'
          ,''
        )
        ,' '
        ,'%'
      )
      ,'%'
    )
union
select
  strftime(Date,'%Y-%m-%d') as SessionDate
  ,'' as StartTime
  ,Name as SessionDescription
  ,Color as Color
  ,'' as JudicialOfficerCode
  ,DisplayOrder as DisplayOrder
  ,strftime(Date,'%U')::int as Week
from
  special_date
;
-- ================================================================================================
-- Get list of unique sessions in each week, sorting
-- and adding a row_num so to have ability to maintain 
-- same sort order.
select
  SessionDescription
  ,Week
  ,row_number() over (order by Week,DisplayOrder,SessionDescription) row_num
from
  (
    select distinct
      SessionDescription
      ,DisplayOrder
      ,Week
    from
      tmp_courtsession
    where
      StartTime = ''
    order by
      Week
      ,DisplayOrder
      ,SessionDescription
  ) z
  order by
    Week
    ,DisplayOrder
    ,SessionDescription
;
-- ================================================================================================
"""
        week_sessions = ddb_conn.execute(sql_qry).fetchall()

        sql_qry = """
-- ================================================================================================
select
  SessionDate
  ,StartTime
  ,SessionDescription
  ,Color
  ,JudicialOfficerCode
  ,DisplayOrder
  ,strftime(SessionDate,'%U')::int as week
from
  tmp_courtsession
order by
  SessionDate
  ,DisplayOrder
  ,StartTime
  ,SessionDescription
;
-- ================================================================================================
"""
        court_session_list = ddb_conn.execute(sql_qry).fetchall()

        date_list = [s[0] for s in court_session_list]
        for dt in date_range_generator(min(date_list),max(date_list)):
            # Skip weekend dates.
            if dt.isoweekday() <= 5:
                date_sessions = [s for s in court_session_list if s[0] == dt and s[1] == '']
                # From sorted list of all the sessions of the given week,
                # clear those that are not in the current date's list. 
                # For those that are, add any needed additional data, such as color.
                for week_session in [ws for ws in week_sessions if ws[1] == int(dt.strftime("%U"))]:
                    # Get matching date session to current week session, if exists.
                    matching_date_session = None
                    try:
                        matching_date_session = next(
                            date_session for date_session in date_sessions if week_session[0] == date_session[2]
                        )
                    except StopIteration:
                        pass
                    # Using SessionDescription
                    if matching_date_session:
                        # Add matching session with week session row_num to maintain order.
                        # [SessionDate,StatrDate,SessionDescription,Color,JudicialOfficerCode,DisplayOrder,week,row_num]
                        sessions.append(list(matching_date_session)+[week_session[2]])
                    else:
                        # Add blank session with week session row_num to maintain order.
                        # [SessionDate,StatrDate='',SessionDescription='',Color='',JudicialOfficerCode='',DisplayOrder=,week,row_num]
                        sessions.append([dt,'','','','',999,week_session[1],week_session[2]])
#    except CatalogException as e:
#        logger.exception(f"\nAn error occurred: {e}")
    except Exception as e:
        logger.exception(f"\nAn error occurred: {e}")
    finally:
        ddb_conn.close()
        logger.info("Succcessfully converted Dataframe to list.")
    sessions += [list(s)+[9999] for s in court_session_list if s[1] != '']
    return sessions
#--------------------------------------------------------------------------------------------------
def apply_abbreviations(sessions: list,yaml_config: dict) -> list:
    abbreviation_keys = yaml_config['data']['abbreviation_keys']
    for session in sessions:
        # Only replace phase with abrev in sessions with non-empty session[1], i.e. StartTime=''.
        if session[1]:
            for phrase,abbrev in abbreviation_keys.items():
                if len(phrase) == 1 and not phrase.isalpha() and not phrase.isdigit() and not phrase.isspace():
                    session[2].replace(phrase,abbrev)
                elif phrase.lower() in session[2].lower():
                    pattern = re.compile(re.escape(phrase), re.IGNORECASE)
                    session[2] = pattern.sub(abbrev, session[2])
    return sessions
#--------------------------------------------------------------------------------------------------
#==================================================================================================
def main(
    config_file: str = typer.Argument(
        f"{Path(__file__).stem}.ini"
        ,help=(
            ""
        )
    )
    ,yaml_config_file: str = typer.Argument(
        f"{Path(__file__).stem}.yaml"
        ,help=(
            ""
        )
    )
    ,calendar_year: int = typer.Argument(
        date.today().year
        ,help=(
            "The year used to generate the session calendar. Defaults to the current year."
            "This overridden by 'calendar_year' setting in the YAML configuration file."
        )
    )
    ,test_yaml_config_file: bool = typer.Option(
        False
        ,help=(
            "Force load of YAML configuration file only."
        )
    )
):
    """
    """
    try:
        # App initialization
        config = configparser.ConfigParser()
        config.read(config_file)
        for k,v in config['Paths'].items():
            path = Path(v)
            if not path.is_dir():
                path.mkdir()
        
        logger = setup_logging(config)
        logger.info(f"Starting: {Path(__file__).stem}")
        
        logger.info(f"Reading YAML configurations from: {yaml_config_file}")
        yaml_config = read_yaml_configuration(yaml_config_file)
        if test_yaml_config_file:
            logger.info(f"YAML configurations loaded successfully.")
            sys.exit()
        #  Data from YAML config
        calendar_year = yaml_config['data'].get('calendar_year',calendar_year)
###        MAX_ROW = yaml_config['constants']['worksheet']['MAX_ROW']
        
        debugging_skip_code = False
        if not debugging_skip_code:
            # Open a new workbook.
            wb = Workbook()
            # Open the active worksheet. This would be the first of the new workbook.
            ws = wb.active
            # Rename the worksheet
            sheet_name = yaml_config['worksheet']['sheet_name']
            ws.title = sheet_name.replace(
                "${calendar_month_name}$"
                ,calendar.month_name[1]
            ).replace(
                "${calendar_year}$"
                ,str(calendar_year)
            )
            
            # For each subkey of ['worksheet'], create the sheet's layout.
            # Add:
            #   Title
            #   Subtitle
            #   Cell Borders
            for k,v in yaml_config['worksheet'].items():
                if not isinstance(v,dict):
                    continue
                logger.info(f"Creating: {k}")
                top_left_cell = v['cell_range']['top_left_cell']
                bottom_right_cell = v['cell_range']['bottom_right_cell']
                min_row, min_col = coordinate_to_tuple(top_left_cell)
                max_row, max_col = coordinate_to_tuple(bottom_right_cell)
                merge_cells = v['cell_range']['merge_cells']
                # Define font
                font = Font(
                    name=v['font']['name'],             # Font type/family
                    size=v['font']['size'],             # Font size (points)
                    bold=v['font']['bold'],             # Optional: Make text bold
                    italic=v['font']['italic'],         # Optional: Make text italic
                    color=yaml_config['constants']['colors'][v['font']['color']]  # Font color (Hex code - FF0000 is Red)
                )
                # Define fill color
                fill_color = PatternFill(
                    start_color=yaml_config['constants']['colors'][v['fill']['start_color']], 
                    end_color=yaml_config['constants']['colors'][v['fill']['end_color']], 
                    fill_type=v['fill']['fill_type']
                )
                # Define border
                border = Border(
                    left=Side(
                        style=v['border']['left']['style'], 
                        color=yaml_config['constants']['colors'][v['border']['left']['color']]
                    ), 
                    right=Side(
                        style=v['border']['right']['style'], 
                        color=yaml_config['constants']['colors'][v['border']['right']['color']]
                    ), 
                    top=Side(
                        style=v['border']['top']['style'], 
                        color=yaml_config['constants']['colors'][v['border']['top']['color']]
                    ), 
                    bottom=Side(
                        style=v['border']['bottom']['style'], 
                        color=yaml_config['constants']['colors'][v['border']['bottom']['color']]
                    ) 
                )
                # Merge cells.
                if merge_cells in ['ByColumn','ByBoth',]:
                    ws.merge_cells(f"{top_left_cell}:{bottom_right_cell}")
                    # Set text
                    ws[top_left_cell] = v['text'][0] if isinstance(v['text'],list) else v['text']
                    # Set font
                    ws[top_left_cell].font = font
                    # Set alignment.
                    ws[top_left_cell].alignment = Alignment(
                        horizontal=v['alignment']['horizontal']
                        ,vertical=v['alignment']['vertical']
                        ,wrapText=True
                    )
                    # Set fill color.
                    ws[top_left_cell].fill = fill_color
                else: # 'ByRow'
                    # Merge cells.
                    for i,col in enumerate(range(min_col,max_col+1)):
                        ws.merge_cells(
                            start_row=min_row
                            ,start_column=col
                            ,end_row=max_row
                            ,end_column=col
                        )
                        # Set text
                        ws.cell(row=min_row, column=col).value = v['text'][i]  if isinstance(v['text'],list) else v['text']
                        # Set font
                        ws.cell(row=min_row, column=col).font = font
                        # Set alignment.
                        ws.cell(row=min_row, column=col).alignment = Alignment(
                            horizontal=v['alignment']['horizontal']
                            ,vertical=v['alignment']['vertical']
                            ,wrapText=True
                        )
                        # Set fill color.
                        ws.cell(row=min_row, column=col).fill = fill_color
                # Add border
                # You must apply the border to ALL cells in the merged range
                # since a single cell's border won't cover the entire merged area.
                for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
                    for cell in row:
                        cell.border = border
                # Set column width
                if v.get('column_width_inches',None) is not None:
                    # Set width of columns
                    cell_width = yaml_config['constants']['worksheet']['EXCEL_CELL_UNIT_PER_INCH'] * v['column_width_inches']
                    _, col_idx_left = coordinate_to_tuple(top_left_cell)
                    _, col_idx_right = coordinate_to_tuple(bottom_right_cell)
                    for col_idx in range(col_idx_left,col_idx_right+1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = cell_width
            
            # Save workbook for debugging.
            xlsx_filename = "wb1.xlsx"
            wb.save(xlsx_filename)
            #sys.exit(0)
            
        if not debugging_skip_code:
            workbook_name = "wb1.xlsx"
            wb = load_workbook(workbook_name)
            # Open the active worksheet. This would be the first of the new workbook.
            ws = wb.active
            # Copy worksheet for remaining months
            yaml_config_ws_title = yaml_config['worksheet']['title']
            title_top_left_cell = yaml_config_ws_title['cell_range']['top_left_cell']
            title = yaml_config_ws_title['text'][0]  if isinstance(yaml_config_ws_title['text'],list) else yaml_config_ws_title['text']
            yaml_config_ws_subtitle = yaml_config['worksheet']['subtitle']
            subtitle_top_left_cell = yaml_config_ws_subtitle['cell_range']['top_left_cell']
            subtitle = yaml_config_ws_subtitle['text'][0]  if isinstance(yaml_config_ws_subtitle['text'],list) else yaml_config_ws_subtitle['text']
            superior_judges = yaml_config['data']['superior_judges']
            judges_count = len(superior_judges)
            if judges_count < 1:
                raise Exception("'superior_judges' are not specified in YAML configuration file.")
            for m in range(2,13):
                # Copy worksheet
                new_ws = wb.copy_worksheet(ws)
                # Rename the new worksheet
                new_ws.title = sheet_name.replace("${calendar_month_name}$",calendar.month_name[m]).replace("${calendar_year}$",str(calendar_year))
                # Change title in sheet
                new_ws[title_top_left_cell] = title.replace("${calendar_month_name}$",calendar.month_name[m]).replace("${calendar_year}$",str(calendar_year))
                # Change subtitle in sheet
                new_ws[subtitle_top_left_cell] = subtitle.replace("${superior_judge}$",superior_judges[m - judges_count*int((m-1)/judges_count) - 1]['name'])
                #
            # Change title in first (January) sheet
            ws[title_top_left_cell] = title.replace("${calendar_month_name}$",calendar.month_name[1]).replace("${calendar_year}$",str(calendar_year))
            # Change subtitle in first (January) sheet
            ws[subtitle_top_left_cell] = subtitle.replace("${superior_judge}$",superior_judges[0]['name'])
            
            # Save workbook for debugging.
            xlsx_filename = "wb2.xlsx"
            wb.save(xlsx_filename)
            #sys.exit(0)
            
        if not debugging_skip_code:
            workbook_name = "wb2.xlsx"
            wb = load_workbook(workbook_name)
            # For each sheet (month) set up month days and placeholders for court sessions.
            # The month day cells are indicated by '${calendar_day}$ placeholder.
            month_day_placeholder = '${calendar_day}$'  ### put in yaml_config
            for month in range(1,13):
                # Select the worksheet by index.
                ws = wb.worksheets[month-1]
                month_num_days = calendar.monthrange(calendar_year, month)[1]
                monthday = 0 # Controls month loop.
                for month_day in range(1,month_num_days+1):
                    if month_day < monthday or get_weekday_number(calendar_year,month,month_day) > 5:
                        continue
                    # Locate top left month day placeholder.
                    row_num = find_first_matching_cell(ws,'A',month_day_placeholder)
                    if row_num is None:
                        raise Exception("Unable to locate top left month day placeholder. {month=}, {month_day=}")
                    # Add another week of rows to the sheet.
                    ws.insert_rows(row_num+2, amount=2)
                    # Copy cells to new rows.
                    for row in range(row_num,row_num+2):
                        for col in range(1,6):
                            copy_cell(ws.cell(row,col),ws.cell(row+2,col))
                    # Get the next month_day that is a workday.
                    if (workday := get_weekday_number(calendar_year,month,month_day)) > 5: # >= Friday
                        continue
                    # For the current week, set workday cell value to month_day if workday is has a month day. Else clear it.
                    monthday = month_day
                    for cell_workday in range(1,6):
                        cell = ws.cell(row_num,cell_workday)
                        if cell_workday < workday or monthday > month_num_days:
                            cell_value = 'Empty'
                        else:
                            cell_value = monthday
                            monthday += 1
                        cell.value = cell_value
                    if monthday+2 > month_num_days: # account for Sat & Sun
                        # Cleanup, Remove unneeded row.
                        row_num += 2
                        ws.delete_rows(row_num, amount=2)
                        break
                #break # month loop
            
            # Save workbook for debugging.
            xlsx_filename = "wb3.xlsx"
            wb.save(xlsx_filename)
            wb = load_workbook(xlsx_filename)
            
        if not debugging_skip_code:
            workbook_name = "wb3.xlsx"
            wb = load_workbook(workbook_name)
            # There are now month day cells that contain the value "Empty". 
            # Locate these cell and set the font color to the same as the cell's fill color
            # so to "hide" the content.
            for month in range(1,13):
                # Select the worksheet by index.
                ws = wb.worksheets[month-1]
                # For each column, locate the cells whose value is "Empty" 
                # and then set its cell font color to the same as its fill color.
                # (NOTE: This had to be done after reloading the workbook from a file
                # because doing this in the previous loop code the cell font and fill 
                # settings seem to be tied when copied.
                for col in range(1,6):
                    # Locate the court session placeholder.
                    row_num = 0
                    while (row_num := find_first_matching_cell_by_col_idx(ws,col,"Empty",start_row=row_num+1)):
                        # Set the font color to same as fill to hide the text.
                        cell = ws.cell(row_num,col)
                        fill = cell.fill
                        font = cell.font
                        cell.font = Font(
                            name=font.name
                            ,size=font.size
                            ,bold=font.bold
                            ,italic=font.italic
                            ,underline=font.underline
                            ,strike=font.strike
                            ,color=fill.fgColor.rgb  
                        )
            # Save workbook for debugging.
            xlsx_filename = "wb3a.xlsx"
            wb.save(xlsx_filename)
            wb = load_workbook(xlsx_filename)
            
        if not debugging_skip_code:
            workbook_name = "wb3a.xlsx"
            wb = load_workbook(workbook_name)
            # Get the courts sessions from Odyssey DB for the calendar year,
            # plus the special_dates.
            court_sessions_df = get_odyssey_court_sessions_by_year(calendar_year,config)
            court_session_list = convert_df_to_list(court_sessions_df,yaml_config)
            court_session_list = apply_abbreviations(court_session_list,yaml_config)
            # For each sheet (month) add court sessions to the month_days.
            # The month day sessions cells are indicated by '${calendar_day}$' placeholder.
            court_session_placeholder = '${court_session}$'  ### put in yaml_config
            for month in range(1,13):
                # Select the worksheet by index.
                ws = wb.worksheets[month-1]
                month_num_days = calendar.monthrange(calendar_year, month)[1]
                for month_day in range(1,month_num_days+1):
                    # Get the workday for the month_day.
                    # Continue, if a weekend.
                    workday = get_weekday_number(calendar_year,month,month_day)
                    if workday > 5:
                        continue
                    # Locate the month in the work_day column.
                    row_num = find_first_matching_cell_by_col_idx(ws,workday,month_day)
                    # Locate the court session placeholder.
                    row_num = find_first_matching_cell_by_col_idx(ws,workday,court_session_placeholder,start_row=row_num)
                    if row_num is None:
                        logger.error(f"{month=},{month_day=},{workday=},{row_num=},{day_sessions=}")
                        wb.save("wb-error.xlsx")
                        breakpoint() # Exception debugging.
                        raise Exception(f"Unable to locate court session placeholder. {month=}, {month_day=}, {workday=}")
                    # Clear row cells leading up to the month_day's workday if they contain court_session_placeholder.
                    for wd in range(1,workday):
                        if ws.cell(row_num,wd).value == court_session_placeholder:
                            ws.cell(row_num,wd).value = None
                    if month_day > month_num_days:
                        break # Break to next month
                    day_sessions = [row for row in court_session_list if row[0] == datetime.strptime(f'{calendar_year}-{month:02d}-{month_day:02d}','%Y-%m-%d').date()]
                    if day_sessions:
                        # For the current month_day, add its court_sessions.
                        day_row_num = 0
                        for day_session in day_sessions:
                            # if next row does not contain court session placeholder,
                            # then add a new row of placeholders.
                            if ws.cell(row_num+day_row_num+1,workday).value != court_session_placeholder:
                                # Add court session another row to the sheet.
                                ws.insert_rows(row_num+day_row_num+1, amount=1)
                                # Copy cells to new rows.
                                # This copies the court placeholders from the previous cells.
                                # This row will have just the court session placeholder in each cell. (Could just set the cell values to the placeholder???)
                                for col in range(1,6):
                                    copy_cell(ws.cell(row_num+day_row_num,col),ws.cell(row_num+day_row_num+1,col))
                            session_description, color_name, judicial_officer = day_session[2],day_session[3],day_session[4]
                            session_cell = ws.cell(row_num+day_row_num,workday)
                            if color_name:
                                new_color = get_yaml_config_color(yaml_config,color_name)
                            else:
                                new_color = get_yaml_config_color(yaml_config,'Black')
                            # Add the color to the value to be used later.
                            # Trying to add the color now, when cells are being inserted and copied
                            # seems to cause the colors to be incorrect in the result.
                            if session_description:
                                session_cell.value = f"{session_description}-[{new_color}]"
                            else:
                                session_cell.value = session_description
                            day_row_num += 1
                        row_num += day_row_num
                    # All day sessions, if any, have been added.
                    # Clear the day's remaining session placeholders
                    while row := find_first_matching_cell_by_col_idx(ws,workday,court_session_placeholder):
                        # row must be same or adjacent to row_num.
                        if row and 0 <= (row - row_num) <= 1:
                            row_num = row
                            ws.cell(row_num,workday).value = None
                        else:
                            break
                    if 1==0 and month_day > 13: # test code to stop early.
                        break # month_day loop
                #break # month loop
                 
            # Save workbook for debugging.
            xlsx_filename = "wb4.xlsx"
            wb.save(xlsx_filename)
            wb = load_workbook(xlsx_filename)
            
        if not debugging_skip_code:
            workbook_name = "wb4.xlsx"
            wb = load_workbook(workbook_name)
            # For each month (sheet):
            #   - Remove all court session placeholders.
            #   - Remove all blank rows.
            #   - Add Border to last row.
            for month in range(1,13):
                # Select the worksheet by index.
                ws = wb.worksheets[month-1]
                # Remove all court session placeholders.
                for row in range(ws.max_row+1,2,-1): #range(1,ws.max_row+2)
                    for col in range(1,6):
                        # Remove all court session placeholders.
                        if ws.cell(row,col).value == court_session_placeholder:
                            ws.cell(row,col).value = ""
                # Remove all blank rows.
                for row in range(ws.max_row+1,2,-1): #range(1,ws.max_row+2)
                    blank_row = True
                    for col in range(1,6):
                        if not (not ws.cell(row,col).value or ws.cell(row,col).value == ""):
                            blank_row = False
                            break
                    # Remove blank rows.
                    if blank_row and not ws.cell(row,1).coordinate in ws.merged_cells:
                        ws.delete_rows(row, amount=1)

            # Save workbook for debugging.
            xlsx_filename = "wb5.xlsx"
            wb.save(xlsx_filename)
            
        if not debugging_skip_code:
            workbook_name = "wb5.xlsx"
            wb = load_workbook(workbook_name)
            # Add Border to last row having data.
            for month in range(1,13):
                # Define border
                cs_border = yaml_config['worksheet']['court_session']['border']
                last_border = Border(
                    left=Side(
                        style=cs_border['left']['style'], 
                        color=yaml_config['constants']['colors'][cs_border['left']['color']]
                    ), 
                    right=Side(
                        style=cs_border['right']['style'], 
                        color=yaml_config['constants']['colors'][cs_border['right']['color']]
                    ), 
                    top=Side(
                        style=cs_border['top']['style'], 
                        color=yaml_config['constants']['colors'][cs_border['top']['color']]
                    ), 
                    bottom=Side(
                        style='thick', 
                        color=yaml_config['constants']['colors'][cs_border['bottom']['color']]
                    ) 
                )
                # Select the worksheet by index.
                ws = wb.worksheets[month-1]
                # From the max_row toward top, find the first row with value.
                border_row = None
                for row in range(ws.max_row+1,2,-1): #range(1,ws.max_row+2)
                    blank_row = True
                    for col in range(1,6):
                        if not (not ws.cell(row,col).value or ws.cell(row,col).value == ""):
                            blank_row = False
                            break
                    if not blank_row:
                        # You must apply the border to ALL cells in the merged range
                        # since a single cell's border won't cover the entire merged area.
                        border_row = row
                        break
                for col in range(1,6): 
                    #ws.cell(border_row,col).font # Simply accessing another style property sometimes forces the update
                    ws.cell(border_row,col).border = last_border

            # Save workbook for debugging.
            xlsx_filename = "wb6.xlsx"
            wb.save(xlsx_filename)
            
        if not debugging_skip_code:
            workbook_name = "wb6.xlsx"
            wb = load_workbook(workbook_name)
            # For each sheet (month) look for adjacent cells in each row that have the same content and
            # if cell.value not blank/None or starts with a number, then merge the cells.
            for month in range(1,13):
                ws = wb.worksheets[month-1]
                for row in range(7,ws.max_row+1):
                    first_cell_in_merge = None
                    last_cell_in_merge = None
                    for col in range(2,6):
                        if ws.cell(row,col).value == "9:00 ST CR - Pleas (D)-[FF00B050]": # test code
                            pass
                        if (
                                ws.cell(row,col).value
                                and ws.cell(row,col).value != "Empty"
                                and str(ws.cell(row,col).value)[0] not in "1234567890" 
                                and ws.cell(row,col).value == ws.cell(row,col-1).value
                            ):
                            if first_cell_in_merge is None:
                                first_cell_in_merge = ws.cell(row,col-1)
                                # Set the font of the first_cell_in_merge.
                                first_cell_in_merge.alignment = Alignment(
                                    horizontal='center'
                                    ,vertical='center'
                                    ,wrapText=True
                                )
                                #first_cell_in_merge.fill = yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                            last_cell_in_merge = ws.cell(row,col)
                        if col == 5 or ws.cell(row,col).value and ws.cell(row,col).value != ws.cell(row,col-1).value:
                            if first_cell_in_merge and last_cell_in_merge:
                                # Merge the cells
                                ws.merge_cells(f"{first_cell_in_merge.coordinate}:{last_cell_in_merge.coordinate}")
                                # Define border
                                border = Border(
                                    left=Side(
                                        style='thick', 
                                        color=get_yaml_config_color(yaml_config,'Black')
                                    ), 
                                    right=Side(
                                        style='thick', 
                                        color=get_yaml_config_color(yaml_config,'Black')
                                    ), 
                                    top=Side(
                                        style='thick', 
                                        color=get_yaml_config_color(yaml_config,'Black')
                                    ), 
                                    bottom=Side(
                                        style='thick', 
                                        color=get_yaml_config_color(yaml_config,'Black')
                                    ) 
                                )
                                # Add border
                                # You must apply the border to ALL cells in the merged range
                                # since a single cell's border won't cover the entire merged area.
                                for cells_in_row in ws.iter_rows(
                                    min_row=first_cell_in_merge.row
                                    ,min_col=first_cell_in_merge.column
                                    ,max_row=last_cell_in_merge.row
                                    ,max_col=last_cell_in_merge.column
                                ):
                                    for cell in cells_in_row:
                                        cell.border = border
                                first_cell_in_merge = None
                                last_cell_in_merge = None

            # Save workbook for debugging.
            xlsx_filename = "wb7.xlsx"
            wb.save(xlsx_filename)
            
        debugging_skip_code = False
        if not debugging_skip_code:
            workbook_name = "wb7.xlsx"
            wb = load_workbook(workbook_name)
            # For each sheet (month) set the color for each court session cell.
            # If cell value does not start with a digit, then set the color as the background 
            # and add thick border.
            # Else, set the color of the font.
            #
            # The color has been embeded it the content of the cell, such as:
            #   NEW YEAR'S DAY-[FFFFFFFF]
            # It is removed.
            extraction_pattern = r'-\[([^\]]+)\]'
            pattern_to_remove = r'-\[.*?\]'
            for month in range(1,13):
                ws = wb.worksheets[month-1]
                for row in range(1,ws.max_row+1):
                    for col in range(1,6):
                        cell = ws.cell(row,col)
                        s = str(cell.value)
                        match = re.search(extraction_pattern,s)
                        result_removed = re.sub(pattern_to_remove,'',s)
                        if match:
                            # Group 1 (the content inside the parentheses) holds the desired substring
                            new_color = match.group(1)
                        else:
                            new_color = None
                        if new_color:
                            font = cell.font
                            if str(cell.value)[0] not in "1234567890": #cell.coordinate in ws.merged_cells:
                                cell.fill = PatternFill(start_color=new_color, end_color=new_color, fill_type='solid')
                                cell.font = Font(
                                    name=font.name
                                    ,size=10 #font.size
                                    ,bold=True
                                    ,italic=font.italic
                                    ,underline=font.underline
                                    ,strike=font.strike
                                    ,color=None  
                                )
                                cell.alignment = Alignment(
                                    horizontal='center'
                                    ,vertical='center'
                                    ,wrapText=True
                                )  ## This should come from yaml_config???
                                border = cell.border
                                cell.border = Border(
                                    left=Side(style='thick', color=border.left.color)
                                    ,right=Side(style='thick', color=border.right.color)
                                    ,top=Side(style='thick', color=border.top.color)
                                    ,bottom=Side(style='thick', color=border.bottom.color)
                                )
                            else:
                                cell.font = Font(
                                    name=font.name
                                    ,size=font.size
                                    ,bold=font.bold
                                    ,italic=font.italic
                                    ,underline=font.underline
                                    ,strike=font.strike
                                    ,color=new_color  # Only this property is changed
                                )
                            cell.value = result_removed

            # Save the workbook.
            workbook_name = "wb8.xlsx"
            logger.info(f"Saving workbook: {workbook_name}")
            wb.save(workbook_name)

        if not debugging_skip_code:
            workbook_name = "wb3.xlsx"
            wb = load_workbook(workbook_name)
            #breakpoint()
            workbook_name = "wb8.xlsx"
            wb = load_workbook(workbook_name)
            # Remove all empty cells where possible.
            # Loop through all cells on each month sheet.
            # If non-merged and empty (value=None or '') then find first non-merged/non-empty/non-month day cell below in same column
            # and Copy/Paste that cell into the empty cell, then empty the copied cell.
            #
            # Finally, remove rows where all columns are empty.
    #####        for month in range(1,2):
    #####            ws = wb.worksheets[month-1]
    #####            for col in range(1,6):
    #####                empty_cell = None
    #####                for row in range(1,ws.max_row+1):
    #####                    cell = ws.cell(row,col)
    #####                    if not cell.coordinate in ws.merged_cells and not cell.value:
    #####                        empty_cell = cell
    #####                        # If current cell is not merged and its value is ''/None (empty), then 
    #####                        # get the next non-merged and non-empty cell.
    #####                        for row_num in range(row+1,ws.max_row+1):
    #####                            cell = ws.cell(row,col)
    #####                            if not cell.coordinate in ws.merged_cells and cell.value:
    #####                        row_num = find_first_matching_cell_by_col_idx(ws,col,'',start_row=row+1)
    #####        # Save the workbook.
    #####        workbook_name = "wb9.xlsx"
    #####        logger.info(f"Saving workbook: {workbook_name}")
    #####        wb.save(workbook_name)

        # Save the workbook.
        workbook_name = f"{Path(__file__).stem}.xlsx"
        logger.info(f"Saving workbook: {workbook_name}")
        wb.save(workbook_name)

    except KeyError as e:
        logger.exception(f"FATAL: Missing required configuration key: {e}")
    except configparser.NoSectionError as e:
        logger.exception(f"FATAL: Configuration is missing required section(s). Details: {e}")
    except KeyboardInterrupt:
        logger.info("Program interrupted by user.")
    except Exception as e:
        # Catch-all block: Logs the exception and terminates.
        logger.exception(f"FATAL: Application terminated due to an unexpected unhandled exception: {e}")
#==================================================================================================
if __name__ == "__main__":
    typer.run(main)
