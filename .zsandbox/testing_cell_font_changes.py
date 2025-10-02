# testing_cell_font_changes.py 20251001

import calendar
import configparser
from datetime import date, datetime, timedelta
import duckdb
from enum import Enum
from icecream import ic
from loguru import logger
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Side, Fill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
import os
import pandas as pd
from pandas import DataFrame # Import the specific class for hinting
from pathlib import Path
import pyodbc
import re
from sqlalchemy import create_engine
import sys
import urllib
import typer
import yaml

xlsx_filename = r"..\wb7.xlsx"
wb = load_workbook(xlsx_filename)

extraction_pattern = r'-\[([^\]]+)\]'
pattern_to_remove = r'-\[.*?\]'
for month in range(1,13):
    ws = wb.worksheets[month-1]
    for row in range(1,ws.max_row+1):
        for col in range(1,6):
            cell = ws.cell(row,col)
            s = str(cell.value)
            match = re.search(extraction_pattern,s)
            result_removed = re.sub(pattern_to_remove,'',s)
            if match:
                # Group 1 (the content inside the parentheses) holds the desired substring
                new_color = match.group(1)
            else:
                new_color = None

            if new_color:
                if cell.coordinate in ws.merged_cells:
                    cell.fill = PatternFill(start_color=new_color, end_color=new_color, fill_type='solid')
                else:
                    font = cell.font
                    cell.font = Font(
                        name=font.name,
                        size=font.size,
                        bold=font.bold,
                        italic=font.italic,
                        underline=font.underline,
                        strike=font.strike,
                        color=new_color  # Only this property is changed
                    )
                cell.value = result_removed
#    print(f"{cell.font=}")

#breakpoint()
xlsx_filename = r"wb6-test.xlsx"
print(xlsx_filename)
wb.save(xlsx_filename)


#                        cell_font = session_cell.font
#                        session_cell.font = Font(
#                            name=cell_font.name,
#                            size=cell_font.size,
#                            bold=cell_font.bold,
#                            italic=cell_font.italic,
#                            underline=cell_font.underline,
#                            strike=cell_font.strike,
#                            color=new_color  # Only this property is changed
#                        )
                        #session_cell.fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
