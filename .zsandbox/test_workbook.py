# test_workbook.py 20250925

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple


border = Border(
    left=Side(
        style='thick', 
        color='FF0000FF'
    ), 
    right=Side(
        style='thick', 
        color='FF0000FF'
    ), 
    top=Side(
        style=None, 
        color='FF0000FF'
    ), 
    bottom=Side(
        style='thick', 
        color='FF0000FF'
    ) 
)

wb = load_workbook("wb5.xlsx")

# Select the worksheet by index.
ws = wb.worksheets[0]

row = 53
col = 1
ws.cell(row,col).value
ws.cell(row,col).border = border

for col in range(1,6):
    ws.cell(row,col).border = border

wb.save("wb5s.xlsx")
