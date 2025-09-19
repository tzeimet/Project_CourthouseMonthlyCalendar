# testing_db_sp_to_xlsx.py 20250918/20250919

"""
Test connecting to and querying a MS SQL database.

Uses the pacakes:
pyodbc

# For Windows Authentication (if you're on a trusted network):
# CONNECTION_STRING = (
#     "DRIVER={ODBC Driver 17 for SQL Server};"
#     "SERVER=YourServerName;"
#     "DATABASE=YourDatabaseName;"
#     "Trusted_Connection=yes;"
# )

"""
import duckdb
from icecream import ic
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
from pandas.tseries.offsets import CustomBusinessDay
from pandas.tseries.holiday import Holiday, CustomHoliday, USMothersDay # Import for defining holidays
import pyodbc
import sys

# Connection details
DRIVER_NAME = 'SQL Server' # Verify your driver name
SERVER_NAME = 'fcvodsysqlprod\\GAFORSYTHPROD'
DATABASE_NAME = 'Justice'
USERNAME = 'YourUsername'
PASSWORD = 'YourPassword'

# Connection string using f-string for clarity
# For Windows Authentication (if you're on a trusted network):
CONNECTION_STRING = (
    f"DRIVER={DRIVER_NAME};"
    f"SERVER={SERVER_NAME};"
    f"DATABASE={DATABASE_NAME};"
    f"Trusted_Connection=yes;"
)
STORED_PROC_NAME = "Justice.fc.sp_getCourtSessionsByYear"
PMONTH = 9

SQL_CALL = f"{{CALL {STORED_PROC_NAME} (?)}}"

# Use a '?' for each parameter the stored procedure accepts.
sql = """\
SET NOCOUNT ON;
EXEC Justice.fc.sp_getCourtSessionsByYear @pMonth=?;
"""

COURT_SESSION_CREATE_TABLE = """
create table court_session
(
    SessionDate date
    ,StartTime time
    ,SessionDescription varchar
    ,CalendarDescription varchar
    ,JudicialOfficerCode varchar
    ,JudicialOfficerDescription varchar
)
;
"""
COURT_SESSION_INSERT = """
insert into court_session
(
    SessionDate
    ,StartTime
    ,SessionDescription
    ,CalendarDescription
    ,JudicialOfficerCode
    ,JudicialOfficerDescription
)
values
(
    $SessionDate
    ,$StartTime
    ,$SessionDescription
    ,$CalendarDescription
    ,$JudicialOfficerCode
    ,$JudicialOfficerDescription
)
"""

sqlconn = None # Initialize sqlconnection

#==================================================================================================
# Spreadsheet Constansts
#==================================================================================================
FONT_COLORs = {
    'JSB': Font(color="FFFF0000", bold=True) # Red
    ,'KMC': Font(color="FFFFA500", bold=True) # Orange
    ,'DLD': Font(color="FF800080", bold=True) # Purple
    ,'JAD': Font(color="FF008000", bold=True) # Green
    ,'SUP': Font(color="FFADD8E6", bold=True) # Light Blue
    ,'TRM': Font(color="FF000000", bold=True) # Black
    ,'JMM': Font(color="FFEE82EE", bold=True) # Violet
    ,'PCS': Font(color="FF0000FF", bold=True) # Blue
    ,'CWW': Font(color="FFA52A2A", bold=True) # Brown
}
#==================================================================================================
try:
    ddb_conn = duckdb.connect(database=':memory:')
    sql_conn = pyodbc.connect(CONNECTION_STRING)
    cursor = sql_conn.cursor()
    
    df = pd.read_sql(
        sql=sql
        ,con=sql_conn
        ,params=(9)
    )

    breakpoint()
    sys.exit()
    
    ddb_conn = duckdb.connect(database=':memory:')
    ddb_conn.execute(COURT_SESSION_CREATE_TABLE)

    # 1. sqlconnect to the database
    sql_conn = pyodbc.connect(CONNECTION_STRING)
    cursor = sql_conn.cursor()
    print("Successfully sqlconnected to the database.")
    
     # 2. Execute the stored procedure
    cursor.execute(SQL_CALL, (9))
    print(f"Executed stored procedure: {STORED_PROC_NAME}")

    # 3. Handle results (if the stored procedure returns data)
    # Use fetchall(), fetchone(), or iterate over the cursor
    
    # Open a new workbook.
    wb = Workbook()
    sheet_nbr = 0
    while True:
        if cursor.description:
            sheet_nbr += 1
            print(cursor.description)
            #input("Pausing...Hit any key to continue.")
            if sheet_nbr == 1:
                # Open the active worksheet. This would be the first of the new workbook.
                ws = wb.active
            else:
                ws = wb.create_sheet()
            ws.title = f"Sheet{sheet_nbr}"
            # Set column headers.
            col_width = []
            session_description_col = 0
            judical_code_col = 0
            header = []
            for col,hdr in enumerate(cursor.description):
                ws.cell(row=1, column=col+1, value=hdr[0])
                col_width.append(len(hdr[0]))
                header.append(hdr[0])
                match hdr[0]:
                    case 'SessionDescription':
                        session_description_col = col+1
                    case 'JudicialOfficerCode':
                        judical_code_col = col+1
            ic(header)
            
            print("\nStored Procedure Results:")
            for row_idx, row in enumerate(cursor.fetchall()):
                values = {}
                for col_idx, value in enumerate(row):
                    values[header[col_idx]] = value
                    print(f"({row_idx+2},{col_idx+1}), {value=}")
                    ws.cell(row=row_idx+2, column=col_idx+1, value=value)
                    if type(value) is str and len(value) > col_width[col_idx]:
                        col_width[col_idx] = len(value)
                ddb_conn.execute(COURT_SESSION_INSERT, values)

            # Set column widths.
            # Excel column widths are based on character count, but often need padding.
            # A factor (e.g., 1.2 or 1.5) and a fixed padding (e.g., +2) are commonly used 
            # to ensure text doesn't look too cramped.
            for col_idx, col_w in enumerate(col_width):
                adjusted_width = (col_w + 2) * 1.2
                adjusted_width = (col_w + 0) * 1.2
                adjusted_width = col_w
                # Assign the calculated width to the column dimension
                ws.column_dimensions[get_column_letter(col_idx+1)].width = adjusted_width
            
        if not cursor.nextset():
            break

    # Save the workbook file
    wb.save("testing_db_sp_to_xls.xlsx")
    
#    # 4. Commit changes (if the stored procedure modifies data)
#    sql_conn.commit()
#    print("\nTransaction committed.")

    # Test the table.
    input("Pausing...Hit any key to continue.")
    results = ddb_conn.sql("SELECT * FROM court_session;").fetchall()
    ic("--- Results ---")
    ic(results)
    
except pyodbc.Error as ex:
    sqlstate = ex.args[0]
    print(f"Database Error: {sqlstate}")
    if sql_conn:
        sql_conn.rollback() # Rollback in case of an error

finally:
    # Close DB connections.
    if ddb_conn:
        ddb_conn.close()
        print("DuckDB connection closed.")

    if sql_conn:
        sql_conn.close()
        print("sqlconnection closed.")
