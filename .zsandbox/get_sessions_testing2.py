# get_sessions_testing2.py 20251002

import calendar
import configparser
import copy
from datetime import date, datetime, timedelta
import duckdb
from enum import Enum
from icecream import ic
from loguru import logger
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Side, Fill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
import os
import pandas as pd
from pandas import DataFrame # Import the specific class for hinting
from pathlib import Path
import pyodbc
import re
from sqlalchemy import create_engine
import sys
import urllib
import typer
import yaml

# ------------------------------------------------------------------------------------------------+
#==============================================================================
# Functions
#==============================================================================
def setup_logging(config: configparser.ConfigParser):
    logs_folder = config['Paths']['logs_folder']
    # console logging
    console_log_level = config['Logging']['console_level']
    # Define a custom format string with color tags for the console output
    console_format = "<light-green>{time:YYYY-MM-DD HH:mm:ss}</light-green> | <level>{level: <8}</level> | <bold><cyan>{function:<20}</cyan>:<cyan>{line:>5}</cyan></bold> - <level>{message}</level>"
    console_format = config['Logging'].get('console_format',console_format)
    # file logging
    file_log_level = config['Logging']['file_level']
    # Define a custom log file path with a dynamic date
    file_log_path = Path(logs_folder) / Path(f"{Path(__file__).stem}"+"_{time:YYYY-MM-DD}.log")
    # Define a simple format string for the log file (no colors needed)
    file_format = "{time:YYYY-MM-DD HH:mm:ss} | {level:<8} | {function:<20}:{line:>5} - {message}"
    file_format = config['Logging'].get('file_format',file_format)
    # Remove the default handler
    logger.remove(0)
    # Add a sink for the console with the colored format
    logger.add(
        sys.stderr, # Use sys.stderr for console output
        level=console_log_level.upper(), 
        format=console_format,
        backtrace=False,
        diagnose=False,
        enqueue=True
    )
    # Add a sink for the log file with the plain format, daily rotation, and retention
    logger.add(
        file_log_path, 
        level=file_log_level.upper(), 
        rotation="1 day", 
        retention="1 week",
        format=file_format,
        backtrace=True,
        diagnose=True,
        enqueue=True
    )
    return logger
# =================================================================================================
# Read applicaytion configuration from a YAML file.
# =================================================================================================
def read_yaml_configuration(yamlFilename: str) -> dict:
    """
    Reads the applicaiton's configuration informaiton from the given filename, yamlFilename. 
    Returns: A tuple (
        dfFinCols: Dictionary containing the Finances DataFrame structure, 
        # This is definition of the DataFrame, with column names and types, that the input CSV data 
        # files will be converted to.
        sourcedata: An array containg a dictionary for each type of input data file
        # Each dictionary defines how to convert its corresponding CSV data file into the finances DataFrame.
    )
    """
    ymldata = None
    with open(yamlFilename, 'r') as file:
        ymldata = yaml.safe_load(file)
    return(ymldata)
# ------------------------------------------------------------------------------------------------+
def get_yaml_config_color(yaml_conf: dict, color_name: str) -> str:
    return yaml_conf['constants']['colors'][color_name]
# ------------------------------------------------------------------------------------------------+
def date_range_generator(start_date: date, end_date: date):
    """Generates a sequence of dates from start_date up to and including end_date."""
    # 1. Calculate the duration (a timedelta object)
    delta = end_date - start_date
    # 2. Iterate using range() for the number of days in the duration
    # delta.days gives the number of days, +1 makes the range inclusive of the end_date.
    for i in range(delta.days + 1):
        # 3. Yield the start_date plus the timedelta for the current iteration (i days)
        yield start_date + timedelta(days=i)
#--------------------------------------------------------------------------------------------------
def get_odyssey_court_sessions_by_year(year: int,config: configparser) -> DataFrame:
    # Construct the ODBC connection string with Trusted_Connection=yes
    odbc_conn_str = (
        f"DRIVER={{{config['SQL']['driver_name']}}};"  # Note the double curly braces for the driver name
        f"SERVER={config['SQL']['server_name']};"
        f"DATABASE={config['SQL']['database_name']};"
        f"Trusted_Connection=yes;"
    )
    # URL-encode the entire ODBC connection string
    params = urllib.parse.quote_plus(odbc_conn_str)
    # Create the Database URI using the mssql+pyodbc dialect
    # The format is: dialect+driver:///?odbc_connect=params
    DB_URI = f"mssql+pyodbc:///?odbc_connect={params}"
    # Create the SQLAlchemy Engine
    engine = create_engine(DB_URI)
    logger.info("SQLAlchemy Engine created successfully using Windows Authentication.")
    # Define SQL stament to execute the stored procedure.
    sql_query = f"""
        SET NOCOUNT ON;
        exec Justice.fc.sp_getCourtSessionsByYear @pMonthOrYear=?;
    """
    # Execute the SQL stored procedure to get all of the Odyssey court sessions for the
    # given year and store the results in a Pandas DataFrame.
    try:
        df = pd.read_sql(
            sql=sql_query,
            con=engine,      # Pass the SQLAlchemy Engine
            params=(year,) 
        )
        df['SessionDate'] = pd.to_datetime(df['SessionDate'])
    except Exception as e:
        logger.exception(f"\nAn error occurred: {e}")
    finally:
        engine.dispose()
    return df
#**************************************************************************************************
# This is the function being tested.
#**************************************************************************************************
def convert_df_to_list(df: DataFrame,yaml_config: dict) -> list:
    court_session_list = []
    sessions = []
    try:
        # Create a connection to an in memory DuckDB DB.
        #ddb_conn = duckdb.connect(database=':memory:')
        ddb_conn = duckdb.connect(database=r'court_calendar2.db')
        #------------------------------------------------------------------------------------------
        # Get the judges from the YAML configuration
        judges = []
        judges = yaml_config['data']['superior_judges'] + yaml_config['data']['state_judges']
        # Create table and insert the judges
        DUCKDB_TABLE_NAME = "judge"
        ddb_conn.sql(f"""
drop table if exists {DUCKDB_TABLE_NAME};
CREATE TABLE {DUCKDB_TABLE_NAME} (
    Name VARCHAR
    ,OysseyCode VARCHAR
    ,Color VARCHAR
 );
        """)
        # Insert the judges
        # DuckDB handles the list insertion automatically when passed as a parameter.
        for record in judges:
            ordered_values = (
                record['name']
                ,record['odyssey_code']
                ,record['color']
            )
            ddb_conn.execute(f"INSERT INTO {DUCKDB_TABLE_NAME} VALUES (?, ?, ?)", ordered_values)
        #------------------------------------------------------------------------------------------
        # Get the Special Dates from the YAML configuration
        special_dates = []
        #breakpoint()
        for sp_dt in yaml_config['data']['special_dates']:
            if sp_dt.get('date',None):
                special_dates.append(sp_dt)
            else:
                for dt in date_range_generator(sp_dt['begin_date'],sp_dt['end_date']):
                    new_sp_dt = {
                        'name': sp_dt['name']
                        ,'date': dt
                        ,'color': sp_dt['color']
                        ,'display_order': sp_dt['display_order']
                    }
                    special_dates.append(new_sp_dt)
        # Create table and insert the special dates.
        DUCKDB_TABLE_NAME = "special_date"
        ddb_conn.sql(f"""
drop table if exists {DUCKDB_TABLE_NAME};
CREATE TABLE {DUCKDB_TABLE_NAME} (
    Name VARCHAR
    ,Date DATE
    ,Color VARCHAR
    ,DisplayOrder int
 );
        """)
        # Insert the special_dates.
        # DuckDB handles the list insertion automatically when passed as a parameter.
        for record in special_dates:
            ordered_values = (
                record['name'] 
                ,record['date'] 
                ,record['color']
                ,record['display_order']
            )
            ddb_conn.execute(f"INSERT INTO {DUCKDB_TABLE_NAME} VALUES (?, ?, ?, ?)", ordered_values)
        #------------------------------------------------------------------------------------------
        # Get the Court Session mappings from the YAML configuration.
        courtsession_mapping_list = yaml_config['data']['sessions_mapping']
        # Create the table schema first
        DUCKDB_TABLE_NAME = "courtsession_mapping"
        ddb_conn.sql(f"""
drop table if exists {DUCKDB_TABLE_NAME};
CREATE TABLE {DUCKDB_TABLE_NAME} (
    OdysseyCourtSession VARCHAR
    ,CalendarFormat VARCHAR
    ,DisplayOrder int
);
        """)
        # Insert the entire list of records using the built-in VALUES clause
        # DuckDB handles the list insertion automatically when passed as a parameter.
        for record in courtsession_mapping_list:
            ordered_values = (
                record['odyssey_name'] 
                ,record['calendar_name'] 
                ,record['display_order']
            )
            ddb_conn.execute(f"INSERT INTO {DUCKDB_TABLE_NAME} VALUES (?, ?, ?)", ordered_values)
        #------------------------------------------------------------------------------------------
        # Create Court Sessions DB table from DataFrame df.
        DUCKDB_TABLE_NAME = 'courtsession'
        ddb_conn.sql(f"""
drop table if exists {DUCKDB_TABLE_NAME};
CREATE TABLE {DUCKDB_TABLE_NAME} AS SELECT * FROM df;
        """
        )
        #------------------------------------------------------------------------------------------
        # Create a list of the court sessions transitioned via the mapping and special_dates
        # and resulting in a the form needed to be displayed in the calendar.
        sql_qry="""
-- ================================================================================================
drop table if exists tmp_courtsession;
CREATE TEMPORARY TABLE tmp_courtsession
(
  SessionDate date
  ,StartTime varchar
  ,SessionDescription varchar
  ,Color varchar
  ,JudicialOfficerCode varchar
  ,DisplayOrder int
  ,Week int
);
insert into
  tmp_courtsession
select
  strftime(SessionDate,'%Y-%m-%d') as SessionDate
  ,case 
    when cs_m.CalendarFormat is not null
    then ''
    else cs.StartTime
  end as StartTime
  ,case 
    when cs_m.CalendarFormat is not null
    then
      replace(
        replace(
          cs_m.CalendarFormat
          ,'$CourtRoom}$'
          ,cs.CourtRoomCode
        )
        ,'${JudicialOfficer}$'
        ,cs.JudicialOfficerCode
      )
    else 
      concat(
        strftime(
          concat(
            '2025-01-01 '
            ,cs.StartTime
          )::datetime
          ,'%-I:%M '
        )
        ,regexp_replace(
          cs.SessionDescription
          ,'\\([A-Z]{3}\\) '
          ,''
        )
        ,' ('
        ,left(reverse(cs.JudicialOfficerCode),1)
        ,')'
      )
  end as SessionDescription
  ,j.Color as Color
  ,cs.JudicialOfficerCode as JudicialOfficerCode
  ,if(cs_m.DisplayOrder is null,3,cs_m.DisplayOrder) as DisplayOrder
  ,strftime(SessionDate,'%U')::int as Week
from
  courtsession cs 
  left outer join judge j
  on
   j.OysseyCode = cs.JudicialOfficerCode
  left outer join courtsession_mapping cs_m 
  on
    cs.SessionDescription ilike concat(
      '%'
      ,replace(
        replace(
          cs_m.OdysseyCourtSession
          ,'-'
          ,''
        )
        ,' '
        ,'%'
      )
      ,'%'
    )
union
select
  strftime(Date,'%Y-%m-%d') as SessionDate
  ,'' as StartTime
  ,Name as SessionDescription
  ,Color as Color
  ,'' as JudicialOfficerCode
  ,DisplayOrder as DisplayOrder
  ,strftime(Date,'%U')::int as Week
from
  special_date
;
-- ================================================================================================
-- Get list of unique sessions in each week, sorting
-- and adding a row_num so to have ability to maintain 
-- same sort order.
select
  SessionDescription
  ,Week
  ,row_number() over (order by Week,DisplayOrder,SessionDescription) row_num
from
  (
    select distinct
      SessionDescription
      ,DisplayOrder
      ,Week
    from
      tmp_courtsession
    where
      StartTime = ''
    order by
      Week
      ,DisplayOrder
      ,SessionDescription
  ) z
  order by
    Week
    ,DisplayOrder
    ,SessionDescription
;
-- ================================================================================================
"""
        week_sessions = ddb_conn.execute(sql_qry).fetchall()

        sql_qry = """
-- ================================================================================================
select
  SessionDate
  ,StartTime
  ,SessionDescription
  ,Color
  ,JudicialOfficerCode
  ,DisplayOrder
  ,strftime(SessionDate,'%U')::int as week
from
  tmp_courtsession
order by
  SessionDate
  ,DisplayOrder
  ,StartTime
  ,SessionDescription
;
-- ================================================================================================
"""
        court_session_list = ddb_conn.execute(sql_qry).fetchall()

        breakpoint()
        date_list = [s[0] for s in court_session_list]
        for dt in date_range_generator(min(date_list),max(date_list)):
            # Skip weekend dates.
            if dt.isoweekday() <= 5:
                date_sessions = [s for s in court_session_list if s[0] == dt and s[1] == '']
                # From sorted list of all the sessions of the given week,
                # clear those that are not in the current date's list. 
                # For those that are, add any needed additional data, such as color.
                for week_session in [ws for ws in week_sessions if ws[1] == int(dt.strftime("%U"))]:
                    # Get matching date session to current week session, if exists.
                    matching_date_session = None
                    try:
                        matching_date_session = next(
                            date_session for date_session in date_sessions if week_session[0] == date_session[2]
                        )
                    except StopIteration:
                        pass
                    # Using SessionDescription
                    if matching_date_session:
                        # Add matching session with week session row_num to maintain order.
                        # [SessionDate,StatrDate,SessionDescription,Color,JudicialOfficerCode,DisplayOrder,week,row_num]
                        sessions.append(list(matching_date_session)+[week_session[2]])
                    else:
                        # Add blank session with week session row_num to maintain order.
                        # [SessionDate,StatrDate='',SessionDescription='',Color='',JudicialOfficerCode='',DisplayOrder=,week,row_num]
                        sessions.append([dt,'','','','',999,week_session[1],week_session[2]])
#    except CatalogException as e:
#        logger.exception(f"\nAn error occurred: {e}")
    except Exception as e:
        logger.exception(f"\nAn error occurred: {e}")
    finally:
        ddb_conn.close()
        logger.info("Succcessfully converted Dataframe to list.")
    sessions += [list(s)+[9999] for s in court_session_list if s[1] != '']
    return sessions
#==============================================================================
def apply_abbreviations(sessions: list,yaml_config: dict) -> list:
    abbreviation_keys = yaml_config['data']['abbreviation_keys']
    breakpoint()
    for session in sessions:
        # Only replace phase with abrev in sessions with non-empty session[1], i.e. StartTime=''.
        if session[1]:
            for phrase,abbrev in abbreviation_keys.items():
                if len(phrase) == 1 and not phrase.isalpha() and not phrase.isdigit() and not phrase.isspace():
                    session[2].replace(phrase,abbrev)
                elif phrase.lower() in session[2].lower():
                    pattern = re.compile(re.escape(phrase), re.IGNORECASE)
                    session[2] = pattern.sub(abbrev, session[2])
    return sessions
#==============================================================================
#**************************************************************************************************
# The above is the function being tested.
#**************************************************************************************************
#--------------------------------------------------------------------------------------------------
if __name__ == "__main__":
### Test code
    config_file = "gen_court_session_calendar.ini"
    yaml_config_file = "gen_court_session_calendar.yaml"
    # App initialization
    config = configparser.ConfigParser()
    config.read(config_file)
    for k,v in config['Paths'].items():
        path = Path(v)
        if not path.is_dir():
            path.mkdir()
    
    logger = setup_logging(config)
    logger.info(f"Starting: {Path(__file__).stem}")
    
    logger.info(f"Reading YAML configurations from: {yaml_config_file}")
    yaml_config = read_yaml_configuration(yaml_config_file)
    
    # Get the courts sessions from Odyssey DB for the calendar year,
    # plus the special_dates.
    calendar_year = yaml_config['data'].get('calendar_year',2025)
    court_sessions_df = get_odyssey_court_sessions_by_year(calendar_year,config)
    court_session_list = convert_df_to_list(court_sessions_df,yaml_config)
    court_session_list = apply_abbreviations(court_session_list,yaml_config)
    breakpoint()
    pass
### Test code
