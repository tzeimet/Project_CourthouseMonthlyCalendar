# test1.py 20250917

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Create sheets in the workbook.
ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
ws3 = wb.create_sheet("Mysheet", -1) # insert at the penultimate position

# Change the name of a sheets
ws1.title = "New Title"

# Get a sheet by name
ws = wb["New Title"]

# Get names of all sheets
print(wb.sheetnames)

# loop through worksheets
for sheet in wb:
    print(sheet.title)
    
 # create copies of worksheets within a single workbook:
source = wb.active
target = wb.copy_worksheet(source)

# Accessing one cell
c = ws['A4']
ws['A4'] = 4
d = ws.cell(row=4, column=2, value=10)

# Accessing many cells
# Ranges of cells
cell_range = ws['A1':'C2']
# Ranges of rows or columns
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
   for cell in row:
       print(cell)

for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

# returns just the cell values:
for row in ws.values:
    for value in row:
        print(value)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)

# assign cell a value:
c.value = 'hello, world'
d.value = 3.14

# Save the file
wb.save("sample.xlsx")

# Save a workbook as a template:
wb = load_workbook('document.xlsx')
wb.template = True
wb.save('document_template.xltx')

# open an existing workbook:
from openpyxl import load_workbook
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)