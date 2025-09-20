# xlsx_formating.py 20250919

from enum import Enum
from icecream import ic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

#==================================================================================================
class ARGBColors(Enum):
     Red            = 'FFFF0000' 
     Blue           = 'FF0000FF'
     Blue_Dark      = 'FF00008B'
     Blue_Light     = 'FFADD8E6'
     Blue_Navy      = 'FF000080'
     Blue_Midnight  = 'FF191970'
     Blue_Royal     = 'FF4169E1'
     Blue_Custom    = 'FF002060'
     Orange         = 'FFFFA500'
     Purple         = 'FF800080'
     Green          = 'FF008000'
     Black          = 'FF000000'
     Violet         = 'FFEE82EE'
     Brown          = 'FFA52A2A'
     White          = 'FFFFFFFF'

title_border = Border(
    left=Side(style='thick', color=ARGBColors.Black.value), 
    right=Side(style='thick', color=ARGBColors.Black.value), 
    top=Side(style='thick', color=ARGBColors.Black.value), 
    bottom=Side(style='thin', color=ARGBColors.White.value)
)
subtitle_border = Border(
    left=Side(style='thick', color=ARGBColors.Black.value), 
    right=Side(style='thick', color=ARGBColors.Black.value), 
    top=Side(style='thin', color=ARGBColors.White.value), 
    bottom=Side(style='thick', color=ARGBColors.Black.value)
)
workday_border = Border(
    left=Side(style='thick', color=ARGBColors.Black.value), 
    right=Side(style='thick', color=ARGBColors.Black.value), 
    top=Side(style='thick', color=ARGBColors.Black.value), 
    bottom=Side(style='thick', color=ARGBColors.Black.value)
)
#==================================================================================================
# Open a new workbook.
wb = Workbook()
# Open the active worksheet. This would be the first of the new workbook.
ws = wb.active
#--------------------------------------------------------------------------------------------------
# Set width of columns TO 1.59"
INCHES_TO_EXCEL_UNIT = 7.2
desired_inches = 1.59 # 0.88
#desired_inches = 2.8 # 1.55
desired_inches = 2.85 # 1.58
desired_inches = 2.87 # 
# Calculate the approximate Excel width
excel_width = desired_inches * INCHES_TO_EXCEL_UNIT
# excel_width will be approximately 11.448
ws.column_dimensions['A'].width = excel_width
ws.column_dimensions['B'].width = excel_width
ws.column_dimensions['C'].width = excel_width
ws.column_dimensions['D'].width = excel_width
ws.column_dimensions['E'].width = excel_width
#--------------------------------------------------------------------------------------------------
# Merge cells for title.
ws.merge_cells('A1:E2')
# Set title text
ws['A1'] = "January 2025 Superior Court Master Calendar"
# Set title font
title_font = Font(
    name='Arial',             # Font type/family
    size=16,                  # Font size (points)
    bold=True,                # Optional: Make text bold
    italic=False,             # Optional: Make text italic
    color=ARGBColors.White.value  # Font color (Hex code - FF0000 is Red)
)
ws['A1'].font = title_font
# Set title alignment.
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
# Set title fill color.
fill_color = PatternFill(start_color=ARGBColors.Blue_Custom.value, end_color=ARGBColors.Blue_Custom.value, fill_type='solid')
ws['A1'].fill = fill_color
# Add title border
# You must apply the border to ALL cells in the merged range
# since a single cell's border won't cover the entire merged area.
for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=5):
    for cell in row:
        cell.border = title_border
#--------------------------------------------------------------------------------------------------
# Merge cells for subtitle.
ws.merge_cells('A3:E3')
# Set subtitle text
ws['A3'] = "Presiding Superior Court Judge: Honorable Jeffrey S. Bagley"
# Set subtitle font
subtitle_font = Font(
    name='Arial',             # Font type/family
    size=9,                   # Font size (points)
    bold=True,                # Optional: Make text bold
    italic=False,             # Optional: Make text italic
    color=ARGBColors.White.value  # Font color (Hex code - FF0000 is Red)
)
ws['A3'].font = subtitle_font
# Set subtitle alignment.
ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
# Set subtitle fill color.
fill_color = PatternFill(start_color=ARGBColors.Blue_Custom.value, end_color=ARGBColors.Blue_Custom.value, fill_type='solid')
ws['A3'].fill = fill_color
# Add subtitle border
# You must apply the border to ALL cells in the merged range
# since a single cell's border won't cover the entire merged area.
for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=5):
    for cell in row:
        cell.border = subtitle_border
#--------------------------------------------------------------------------------------------------
# Merge cells for workday
start_r = 4
for start_c in range(1,6):
    ws.merge_cells(
        start_row=start_r,
        start_column=start_c,
        end_row=start_r+1,
        end_column=start_c
    )
    workday = ''
    match start_c:
        case 1:
            workday = 'MONDAY'
        case 2:
            workday = 'TUESDAY'
        case 3:
            workday = 'WEDNESDAY'
        case 4:
            workday = 'THURSDAY'
        case 5:
            workday = 'FRIDAY'
    ws.cell(row=start_r, column=start_c).value = workday
    # Set workday alignment
    ws.cell(row=start_r, column=start_c).alignment = Alignment(horizontal='center', vertical='center')
    # Set workday font
    subtitle_font = Font(
        name='Calibri',             # Font type/family
        size=10,                   # Font size (points)
        bold=True,                # Optional: Make text bold
        italic=False,             # Optional: Make text italic
        color=ARGBColors.Black.value  # Font color (Hex code - FF0000 is Red)
    )
    ws.cell(row=start_r, column=start_c).font = subtitle_font
# Add workday border
# You must apply the border to ALL cells in the merged range
# since a single cell's border won't cover the entire merged area.
for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=5):
    for cell in row:
        cell.border = workday_border
            
#--------------------------------------------------------------------------------------------------
# Save.
wb.save('xlsx_formating.xlsx')

